from openpyxl import load_workbook
from openpyxl.styles import Font

file = "L:\My Documents\Desktop\output.xlsx"
wb = load_workbook(filename=file)
ws = wb['Worksheet Name']
bolded = Font(bold=True)

# Enumerate the cells in the second row
for cell in ws["1:1"]:
    cell.font = bolded

for cell in ws["A:A"]:
    cell.font = bolded
    
wb.save(filename=file)